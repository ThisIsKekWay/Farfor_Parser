import json
import requests
from bs4 import BeautifulSoup
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
import os


def get_url(url):
    while True:
        response = requests.get(f'{url}')
        if response.status_code == 200:
            return response.text
        else:
            print('Сервис временно недоступен')
            break


def cities():
    with open('cities.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    all_cities = {}
    for city in data:
        all_cities[city['name']] = city['domain']
    return all_cities


def get_categories(url):
    soup = BeautifulSoup(get_url(url), 'html.parser')
    data = soup.body
    cats = data.find_all('a', {'class': 'categories-item'})
    categories = {}
    for cat in cats:
        categories[cat.text] = cat.get('href')
    return categories


def delete_nonnumeric(st):
    return str(''.join(i for i in st if i.isdigit()))


def get_items(url, categories):
    main_data = {}
    for cat_name, category in categories.items():
        soup = BeautifulSoup(get_url(url + category), 'html.parser')
        data = soup.body
        items = data.select('.product.product--main-desktop')
        # items = data.find_all('div', {'class': 'product product--main-desktop'})
        data = {}
        for item in items:
            name = item.find('div', {'class': 'product__content-title'}).text
            price = delete_nonnumeric(item.find('div', {'class': 'product__content-price'}).text[0:4])
            weight = delete_nonnumeric(item.find('div', {'class': 'product__content-weight'}).text[0:4])
            if item.find('div', {'class': 'product__image-quantity'}) is None:
                quantity = ""
            else:
                quantity = delete_nonnumeric(item.find('div', {'class': 'product__image-quantity'}).text[0:4])
            data[name] = {'price': price, 'weight': weight, 'quantity': quantity}
        main_data[cat_name] = data
    return main_data


def make_it_excel(data, city):
    if os.path.exists(f'Фарфор {city}.xlsx'):
        wb = load_workbook(f'Фарфор {city}.xlsx')
    else:
        wb = Workbook()
    ws = wb.active

    if ws.max_row > 1:
        wb.create_sheet(f'{datetime.datetime.now().date()}')
        ws = wb[f'{datetime.datetime.now().date()}']
    else:
        ws.title = f'{datetime.datetime.now().date()}'
    ws['A1'] = 'Категория'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    color_scale_rule = ColorScaleRule(start_type='min', start_color='00FF00', end_type='max', end_color='FF0000')

    row = 2
    for cat_name, items in data.items():
        first_row = row + 1
        ws[f'A{row}'] = cat_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = 'Название'
        ws[f'C{row}'] = 'Цена'
        ws[f'D{row}'] = 'Вес'
        if any(details['quantity'] for details in items.values() if details.get('quantity')):
            ws[f'E{row}'] = 'Кол-во'
            ws[f'F{row}'] = 'Стоимость/Вес'
            ws[f'G{row}'] = 'Стоимость/Кол-во'
            row = ws.max_row + 1
            for item, details in items.items():
                ws[f'B{row}'] = item
                ws[f'C{row}'] = details['price']
                ws[f'D{row}'] = details['weight']
                if details.get('quantity'):
                    ws[f'E{row}'] = details['quantity']
                    ws[f'F{row}'] = f"=C{row}/D{row}"
                    ws[f'G{row}'] = f"=C{row}/E{row}"
                elif '=' in ws[f'F{row - 1}'].value:
                    ws[f'F{row}'] = f"=C{row}/D{row}"
                else:
                    ws[f'E{row}'] = f"=C{row}/D{row}"
                row = ws.max_row + 1
            ws.conditional_formatting.add(f'F{first_row}:F{ws.max_row}', color_scale_rule)
            ws.conditional_formatting.add(f'G{first_row}:G{ws.max_row}', color_scale_rule)
        else:
            ws[f'E{row}'] = 'Стоимость/Вес'
            row = ws.max_row + 1
            for item, details in items.items():
                ws[f'B{row}'] = item
                ws[f'C{row}'] = details['price']
                ws[f'D{row}'] = details['weight']
                ws[f'E{row}'] = f"=C{row}/D{row}"
                row = ws.max_row + 1
            if first_row < ws.max_row:
                ws.conditional_formatting.add(f'E{first_row}:E{ws.max_row}', color_scale_rule)
        row += 1
    wb.save(f'Фарфор {city}.xlsx')


def make_it_json(data, city):
    with open(f'{city} {datetime.datetime.now().date()}.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def choose_ur_city():
    for i, city in enumerate(cities.keys()):
        print(f'{i + 1}. {city}')

    print('Введите номер города:\n')
    return int(input())


if __name__ == '__main__':
    cities = cities()
    city = choose_ur_city()
    ur_city = list(cities.keys())[city - 1]
    url = f'https://{cities[list(cities.keys())[city - 1]]}'
    print(f'Выбранный город: {ur_city}. Подключаюсь к домену {url}...')
    cats = get_categories(url)
    print('Категории получены. Собираю данные о товарах...')
    data = get_items(url, cats)
    print('Данные собраны. Выберите формат экспорта:\n 1. Excel\n 2. JSON')
    ans = input()
    while ans != '1' and ans != '2':
        print('Неверная команда')
        ans = input()
    if ans == '1':
        make_it_excel(data, ur_city)
    elif ans == '2':
        make_it_json(data, ur_city)
    else:
        print('Неверная команда')
